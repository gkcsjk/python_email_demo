import json
import base64
import mimetypes
import yaml
import os
import re
import time
from openpyxl import load_workbook
from requests_oauthlib import OAuth2Session


# This is necessary for testing with non-HTTPS localhost
# Remove this if deploying to production
os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'

# This is necessary because Azure does not guarantee
# to return scopes in the same case and order as requested
os.environ['OAUTHLIB_RELAX_TOKEN_SCOPE'] = '1'
os.environ['OAUTHLIB_IGNORE_SCOPE_CHANGE'] = '1'

# Load the oauth_settings.yml file
stream = open('oauth_settings.yml', 'r')
settings = yaml.load(stream, yaml.SafeLoader)
authorize_url = '{0}{1}'.format(settings['authority'], settings['authorize_endpoint'])
token_url = '{0}{1}'.format(settings['authority'], settings['token_endpoint'])
graph_url = 'https://graph.microsoft.com/v1.0'

# useful links
# Microsoft Graph Explorer: https://developer.microsoft.com/en-us/graph/graph-explorer/preview
# Microsoft Graph sendMain doc: https://docs.microsoft.com/en-us/graph/api/user-sendmail?view=graph-rest-1.0&tabs=http
# Microsoft Azure Portal: https://portal.azure.com/#home


def initialize_context(request):
    context = {}

    # Check for any errors in the session
    error = request.session.pop('flash_error', None)

    if error is not None:
        context['errors'] = []
        context['errors'].append(error)

    # Check for user in the session
    context['user'] = request.session.get('user', {'is_authenticated': False})
    return context


def get_sign_in_url():
    o_auth = OAuth2Session(settings['app_id'], scope=settings['scopes'], redirect_uri=settings['redirect'])
    sign_url, state = o_auth.authorization_url(authorize_url, prompt='login')
    return sign_url, state


def get_token_from_code(callback_url, expected_state):
    o_auth = OAuth2Session(
        settings['app_id'],
        state=expected_state,
        scope=settings['scopes'],
        redirect_uri=settings['redirect']
    )

    token = o_auth.fetch_token(token_url, client_secret=settings['app_secret'], authorization_response=callback_url)
    return token


def store_token(request, token):
    request.session['oauth_token'] = token


def store_user(request, user):
    request.session['user'] = {
        'is_authenticated': True,
        'name': user['displayName'],
        'email': user['mail'] if (user['mail'] is not None) else user['userPrincipalName']
    }


def get_token(request):
    token = request.session['oauth_token']
    if token is not None:
        # Check expiration
        now = time.time()
        # Subtract 5 minutes from expiration to account for clock skew
        expire_time = token['expires_at'] - 300
        if now >= expire_time:
            # Refresh the token
            o_auth = OAuth2Session(
                settings['app_id'],
                token=token,
                scope=settings['scopes'],
                redirect_uri=settings['redirect']
            )

            refresh_params = {
                'client_id': settings['app_id'],
                'client_secret': settings['app_secret'],
            }
            new_token = o_auth.refresh_token(token_url, **refresh_params)

            # Save new token
            store_token(request, new_token)

            # Return new access token
            return new_token

        else:
            # Token still valid, just return it
            return token


def remove_user_and_token(request):
    if 'oauth_token' in request.session:
        del request.session['oauth_token']

    if 'user' in request.session:
        del request.session['user']


def get_user(token):
    graph_client = OAuth2Session(token=token)
    # Send GET to /me
    user = graph_client.get('{0}/me'.format(graph_url))
    # Return the JSON result
    return user.json()


def get_clients_from_file():
    clients_list = []
    wb = load_workbook(filename='clients.xlsx', data_only=True)
    clients = wb['Sheet1'].iter_rows(min_row=2, values_only=True)
    for client in clients:
        client_dict = {
            'name': client[0],
            'email': client[1],
            'if_os': client[2],
        }
        clients_list.append(client_dict)

    return clients_list


def check_valid_email(email):
    if email is None:
        return False
    else:
        if not re.fullmatch(r"[^@]+@[^@]+\.[^@]+", email):
            return False
        else:
            return True


def add_attachment(filename):
    b64_content = base64.b64encode(open(filename, 'rb').read())
    mime_type = mimetypes.guess_type(filename)[0]
    mime_type = mime_type if mime_type else ''
    attachment = {
        "@odata.type": "#microsoft.graph.fileAttachment",
        "name": "attachment.docx",
        "contentType": mime_type,
        "contentBytes": b64_content.decode('utf-8')
    }
    return attachment


def send_emails(token):
    graph_client = OAuth2Session(token=token)
    clients = get_clients_from_file()
    for client in clients:
        if check_valid_email(client['email']) and client['if_os'] == 'OS':
            data = {
                'message': {
                    "subject": "Welcome on board!",
                    "body": {
                        "contentType": "Text",
                        "content": "Dear " + client['name'] + "\n\nPlease find the attached file."
                    },
                    "toRecipients": [
                        {
                            "emailAddress": {
                                "address": client['email']
                            }
                        }
                    ],
                    "attachments": [
                        add_attachment('attachment.docx')
                    ]
                }
            }

            graph_client.headers = {
                'Content-type': 'application/json'
            }
            data = json.dumps(data).encode("utf-8")
            res = graph_client.post(
                '{0}/me/sendMail'.format(graph_url),
                data=data,
            )
            if res.status_code == 202:
                print('Success: ' + client['email'])
            else:
                print('Error: ' + client['email'] + " - " + res.json())




